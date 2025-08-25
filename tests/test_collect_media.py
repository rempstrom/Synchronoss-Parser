import importlib.util
from pathlib import Path
from fractions import Fraction

import pytest

pytest.importorskip("openpyxl")
pytest.importorskip("PIL")

from openpyxl import load_workbook


def load_module():
    module_path = Path(__file__).resolve().parents[1] / "scripts" / "collect_media.py"
    spec = importlib.util.spec_from_file_location("collect_media", module_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def test_write_excel_handles_rational(tmp_path):
    collect_media = load_module()

    records = [{
        "File Name": "a.jpg",
        "Date": "2021-01-01",
        "Device": "d",
        "MD5": "0" * 32,
        "ExposureTime": Fraction(1, 2),
        "FNumber": [Fraction(5, 2), Fraction(1, 1)],
    }]
    exif_keys = ["ExposureTime", "FNumber"]

    for k, v in list(records[0].items()):
        records[0][k] = collect_media.normalize_exif_value(v)

    collect_media.LOGFILE = tmp_path / "out.xlsx"
    collect_media.write_excel(records, exif_keys)

    wb = load_workbook(collect_media.LOGFILE)
    ws = wb.active

    assert ws.cell(row=2, column=5).value == 0.5
    assert ws.cell(row=2, column=6).value == "2.5, 1.0"
