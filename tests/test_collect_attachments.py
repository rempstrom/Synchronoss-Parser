import importlib
import hashlib
import sys
from pathlib import Path

from PIL import Image
from openpyxl import load_workbook


def load_module():
    project_root = Path(__file__).resolve().parents[1]
    sys.path.append(str(project_root))
    return importlib.import_module("synchronoss_parser.collect_attachments")


def test_collect_attachments_copies_files_and_logs_metadata(tmp_path):
    collect_attachments = load_module()

    messages_dir = tmp_path / "messages"
    attachments_dir = messages_dir / "attachments" / "mms" / "in" / "2024-01-01"
    attachments_dir.mkdir(parents=True)

    img1 = Image.new("RGB", (10, 10), color="red")
    exif = img1.getexif()
    exif[274] = 1  # Orientation tag
    img1_path = attachments_dir / "photo1.jpg"
    img1.save(img1_path, exif=exif)

    img2 = Image.new("RGB", (10, 10), color="blue")
    img2_path = attachments_dir / "photo2.png"
    img2.save(img2_path)

    csv_content = (
        "Date,Type,Direction,Attachments,Body,Sender,Recipients,\"Message ID\"\n"
        "2024-01-01T00:00:00Z,mms,in,photo1.jpg|photo2.png,Hi,Alice,Bob,id1\n"
    )
    (messages_dir / "20240101.csv").write_text(csv_content)

    compiled = tmp_path / "Compiled Attachments"
    records, exif_keys = collect_attachments.collect_attachments(messages_dir / "attachments", compiled)
    logfile = compiled / "log.xlsx"
    collect_attachments.write_excel(records, exif_keys, logfile)

    dest_files = {f.name for f in compiled.iterdir() if f.is_file() and f.name != "log.xlsx"}
    assert dest_files == {"photo1.jpg", "photo2.png"}

    md5_1 = hashlib.md5(img1_path.read_bytes()).hexdigest()
    md5_2 = hashlib.md5(img2_path.read_bytes()).hexdigest()

    wb = load_workbook(logfile)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    assert headers[:5] == ["File Name", "Date", "Sender", "Recipient", "MD5"]
    assert "Orientation" in headers
    col_map = {h: i + 1 for i, h in enumerate(headers)}

    rows = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, r))
        rows[row_data["File Name"]] = row_data

    assert set(rows.keys()) == {"photo1.jpg", "photo2.png"}

    row1 = rows["photo1.jpg"]
    assert row1["Date"] == "2024-01-01T00:00:00Z"
    assert row1["Sender"] == "Alice"
    assert row1["Recipient"] == "Bob"
    assert row1["MD5"] == md5_1
    assert row1["Orientation"] == 1

    row2 = rows["photo2.png"]
    assert row2["MD5"] == md5_2
    assert row2["Sender"] == "Alice"
    assert row2["Recipient"] == "Bob"
    assert row2.get("Orientation") in ("", None)


def test_duplicate_filenames_different_contexts(tmp_path):
    collect_attachments = load_module()

    messages_dir = tmp_path / "messages"
    messages_dir.mkdir()

    # First attachment
    dir1 = messages_dir / "attachments" / "mms" / "in" / "2024-01-01"
    dir1.mkdir(parents=True)
    img1 = Image.new("RGB", (10, 10), color="red")
    img1_path = dir1 / "sample.png"
    img1.save(img1_path)

    # Second attachment with same filename in different folder
    dir2 = messages_dir / "attachments" / "sms" / "out" / "2024-01-02"
    dir2.mkdir(parents=True)
    img2 = Image.new("RGB", (10, 10), color="blue")
    img2_path = dir2 / "sample.png"
    img2.save(img2_path)

    csv1 = (
        "Date,Type,Direction,Attachments,Body,Sender,Recipients,\"Message ID\"\n"
        "2024-01-01T00:00:00Z,mms,in,sample.png,Hi,Alice,Bob,id1\n"
    )
    csv2 = (
        "Date,Type,Direction,Attachments,Body,Sender,Recipients,\"Message ID\"\n"
        "2024-01-02T00:00:00Z,sms,out,sample.png,Bye,Bob,Alice,id2\n"
    )
    (messages_dir / "20240101.csv").write_text(csv1)
    (messages_dir / "20240102.csv").write_text(csv2)

    compiled = tmp_path / "Compiled Attachments"
    records, _ = collect_attachments.collect_attachments(messages_dir / "attachments", compiled)

    names = {r["File Name"] for r in records}
    assert names == {"sample.png", "sample_1.png"}

    senders = {r["Sender"] for r in records}
    assert senders == {"Alice", "Bob"}

    assert len(list(compiled.glob("*.png"))) == 2
